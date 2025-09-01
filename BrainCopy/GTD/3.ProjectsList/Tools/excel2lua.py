

'''
Author      : babybus
Date        : 2025-06-23 10:15:04
LastEditors : guobiao
LastEditTime: 2025-06-23 10:51:04
FilePath    : excel2lua.py
Description : 

Copyright (c) 2011-2025  baby-bus.com. 


使用说明: 
1.安装openpyxl
· pip3 install openpyxl ·
2.运行
`python excel2lua.py ($表格文件路径) ($lua文件路径)`

注意📢：
表格第一列为table-key，第一行为value-key
'''
    
import os
import sys
import json
import openpyxl
from openpyxl import load_workbook

def escape_chinese(text):
    """
    将中文字符转换为Unicode转义序列
    :param text: 要处理的文本
    :return: 处理后的文本
    """
    if not isinstance(text, str):
        return text
    
    result = []
    for char in text:
        # 如果是中文字符（根据Unicode范围判断）
        if '\u4e00' <= char <= '\u9fff':
            # 转换为Unicode转义序列，格式为\uXXXX
            result.append(f"\\u{ord(char):04x}")
        else:
            result.append(char)
    return ''.join(result)

def excel_to_lua(excel_path, lua_path, sheet_name=None, escape_chinese_chars=True):
    """
    将Excel文件转换为Lua表
    :param excel_path: Excel文件路径
    :param lua_path: 输出的Lua文件路径
    :param sheet_name: 指定工作表名称，默认为第一个工作表
    :param escape_chinese_chars: 是否转义中文字符，默认为True
    """
    try:
        # 加载Excel文件
        wb = load_workbook(excel_path, data_only=True)
        
        # 获取工作表
        if sheet_name:
            sheet = wb[sheet_name]
        else:
            sheet = wb.active
        
        # 获取表头
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        
        # 准备数据
        lua_table = {}
        
        for row in sheet.iter_rows(min_row=2):
            row_data = {}
            for idx, cell in enumerate(row):
                header = headers[idx]
                if header is None:
                    continue
                
                value = cell.value
                # 处理空值
                if value is None:
                    value = ""
                # 处理布尔值
                elif isinstance(value, bool):
                    value = str(value).lower()
                # 处理数字
                elif isinstance(value, (int, float)):
                    # 如果是整数且没有小数部分，保持为整数
                    if isinstance(value, float) and value.is_integer():
                        value = int(value)
                # 处理字符串中的中文字符
                elif isinstance(value, str) and escape_chinese_chars:
                    value = escape_chinese(value)
                
                row_data[header] = value
            
            # 使用第一列作为键（假设第一列是ID）
            if len(row_data) > 0:
                first_key = list(row_data.keys())[0]
                lua_table[row_data[first_key]] = row_data
        
        # 生成Lua代码
        lua_code = f"local data = {{\n"
        
        for key, value in lua_table.items():
            # 处理键的中文转义
            if isinstance(key, str) and escape_chinese_chars:
                key_str = f'"{escape_chinese(key)}"'
            else:
                key_str = json.dumps(key)
            
            lua_code += f"    [{key_str}] = {{"
            for k, v in value.items():
                # 处理键名中的中文转义
                if isinstance(k, str) and escape_chinese_chars:
                    k = escape_chinese(k)
                
                # 根据值的类型决定如何格式化
                if isinstance(v, str):
                    if v.lower() in ('true', 'false'):
                        lua_code += f" {k} = {v.lower()},"
                    else:
                        # 已经在前面的步骤处理过中文转义
                        lua_code += f" {k} = \"{v}\","
                elif isinstance(v, (int, float)):
                    lua_code += f" {k} = {v},"
                else:
                    lua_code += f" {k} = {json.dumps(v)},"
            lua_code += " },\n"
        
        lua_code += "}\n\nreturn data"
        
        # 写入Lua文件
        with open(lua_path, 'w', encoding='utf-8') as f:
            f.write(lua_code)
        
        print(f"成功将 {excel_path} 转换为 {lua_path}")
    
    except Exception as e:
        print(f"转换失败: {str(e)}")

if __name__ == "__main__":
    # 调用转换函数
    excel_to_lua(sys.argv[1], sys.argv[2], escape_chinese_chars=False)